# -*- coding: utf-8 -*-
"""
Author: Andrea Mastrangelo

Last release 02/11/2016

File manipolation library
"""

import os
from openpyxl import load_workbook, Workbook
from decimal import *


def readXls(nameFile, sheet, column, header, limit=0):
    """
    Read values from excel 2010 files column
    """
    a,result=[],[]
    wb = load_workbook(filename = nameFile, read_only=False, data_only=True)
    #data_only to read current values of cells and not formulas
    sheet_ranges = wb[sheet]#worksheet
    
    """
    pay attention with the upgrade of openpyxl 2.4.0 ReadOnlyWS has no more column attribute
    """
    f=sheet_ranges[column]#the first cell with columns is 0-0
    for i in range (len(f)-header):
        a.append(f[i+header].value)
    if limit>0:
        a=a[:limit]
    for item in a:
        if not(item==None):
            result.append(item)
    return result
    
def writeXls(nameFile, h, sheetName="Foglio1"):
    """
    Write h values to excel file
    now it will overwrite an already existing file
    TODO: generalize function
    """
    if fileExist(nameFile):
        wb=load_workbook(filename=nameFile)#load file
    else:
        wb=Workbook()#create new workbook
    ws = wb.create_sheet(title=sheetName)#create new sheet
    ws['A1']="range"
    ws['B1']="cycles"
    ws['C1']="mean"
    ws['D1']="sigma_max"
    ws['E1']="sigma_min"
    ws['F1']="R"
    i=2
    for item in h:
        for value in item[1]:
            ws['A'+str(i)]=item[0]#range
            ws['B'+str(i)]=value[0]#cycles
            ws['C'+str(i)]=value[1]#mean
            s_max=value[1]+item[0]/2
            s_min=value[1]-item[0]/2
            ws['D'+str(i)]=s_max#s_max
            ws['E'+str(i)]=s_min#s_min
            ws['F'+str(i)]=value[2]#R
            i=i+1
    wb.save(nameFile)#save file
    
    
def fileExist(file):
    """
    check if file already exists
    INPUT: file path+name
    OUTPUT: true if exists or false if not
    """
    return os.path.isfile(file)

